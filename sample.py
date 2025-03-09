from consts import const

import os
import openpyxl as excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 現在位置
current_row = const.START_ROW
current_column = const.START_COLUMN

# エクセルの保存先
create_file = os.path.join(os.getcwd(),'create','test.xlsx')
print(create_file)

# ブックの新規作成
# 既にブックが存在する場合は、削除してから新規作成
if os.path.exists(create_file):
    os.remove(create_file)

wb = excel.Workbook()
wb.save(create_file)

# 作成したブックの読み込み
wb = excel.load_workbook(create_file)

# 操作対象のシートを指定
ws = wb.worksheets[0]
# シート名の変更
ws.title = '動作確認用'

loop_count = 0
loop_end = len(const.INPUT_LAYPUT)
for index,item in enumerate(const.INPUT_LAYPUT):
    # 結合の末尾のカラム位置
    end_column_value = current_column+item-1
    # セルに値を代入（タイトル）
    ws.cell(row=current_row,column=current_column,value='メイン' + str(index))
    # セル結合（タイトル）
    ws.merge_cells(start_row=current_row, start_column=current_column,end_row=current_row,end_column=end_column_value)
    # 結合後のセルのスタイルを整える
    ws.cell(current_row,current_column).alignment = Alignment(horizontal='center')
    ws.cell(current_row,current_column).fill = const.CELL_FILL_COLOR
    for column_no in range(current_column, end_column_value+1):
        ws.cell(current_row, column_no).border = Border(left=const.BLACK_THIN, right=const.BLACK_THIN,
                    top=const.BLACK_THIN, bottom=const.BLACK_THIN)
    # 次の列へ
    current_column += item

# サブタイトルの編集に遷移する前に列数をリセット
current_column = const.START_COLUMN
# 次の行（サブタイトル）の編集へ
current_row +=1

for index,item in enumerate(const.INPUT_LAYPUT):       
    for i in range(0,item):
        ws.cell(row=current_row,column=current_column,value='サブ' + str(i))
        # セルのスタイルを整える
        ws.cell(current_row,current_column).alignment = Alignment(horizontal='center')
        ws.cell(current_row,current_column).fill = const.CELL_FILL_COLOR
        for column_no in range(current_column, current_column+1):
            ws.cell(current_row, column_no).border = Border(left=const.BLACK_THIN, right=const.BLACK_THIN,
                        top=const.BLACK_THIN, bottom=const.BLACK_THIN)
        # 次の列へ
        current_column +=1
    

# フォントの指定（シート全体）
font = Font(name=const.FONT_NAME)
for row in ws:
    for cell in row:
        ws[cell.coordinate].font = font

# 上書き保存
wb.save(create_file)