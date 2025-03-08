from consts import const

import os
import openpyxl as excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# エクセルの保存先
create_file = os.path.join(os.getcwd(),'create','sample.xlsx')
print(create_file)

# ブックの新規作成
# 既にブックが存在する場合は、削除してから新規作成
if os.path.exists(create_file):
    os.remove(create_file)

wb = excel.Workbook()
wb.save(create_file)

# 作成したブックの読み込み
wb = excel.load_workbook(create_file)

# 操作対象のシートを指定
ws = wb.worksheets[0]
# セルに値を代入
ws["A1"] = "sample"
# シート名の変更
ws.title = "sample_name"

# ------------------------------------------------------------------------------------
# セル結合1
# 結合するセルの先頭に文字を入れておいて結合を実施する
# 他のセルに入っていた場合は結合すると消える
ws['A2'] = '結合します'
ws.merge_cells('A2:B2')

# セル結合2
ws.cell(row=3,column=1,value='結合2')
ws.merge_cells(start_row=3, start_column=1,end_row=3,end_column=3)

# 結合したセルの文字を中央寄せにする
ws.cell(2,1).alignment = Alignment(horizontal='center')
ws.cell(3,1).alignment = Alignment(horizontal='center')

# 結合したセルの塗りつぶし
ws.cell(3,1).fill = PatternFill(fgColor='C6E0B4', fill_type='solid')

# 結合したセルの罫線
black_thin = Side(color='000000', border_style='thin')
for column_no in range(1, 4):
    ws.cell(3, column_no).border = Border(left=black_thin, right=black_thin,
                top=black_thin, bottom=black_thin)
    
# 結合したセルへの操作は基本的に左上のセルを指定すれば良い！

# ------------------------------------------------------------------------------------

# フォントの指定（シート全体）
font = Font(name='Meiryo UI')
for row in ws:
    for cell in row:
        ws[cell.coordinate].font = font

# 上書き保存
wb.save(create_file)